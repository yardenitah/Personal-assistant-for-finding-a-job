# /Users/yrdnqldrwn/Desktop/SOFTWARE/PayChatm/Info_aboutCVsubmitted/DBManager.py
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill
from pymongo import MongoClient
from dotenv import load_dotenv
import os
import datetime
from openpyxl import Workbook
import matplotlib.pyplot as plt


class DBManager:
    def __init__(self):
        load_dotenv()
        client = MongoClient(os.getenv("DB_URI"))
        self.db = client["JobsDB"]
        self.collection = self.db["MyJobs"]

    def add_job_toDB(self, company_name, link, title, cv_version, hr_or_website):
        date_sent = datetime.datetime.today().strftime('%Y-%m-%d')
        job_entry = {
            "company_name": company_name,
            "job_link": link,
            "job_title": title,
            "cv_version": cv_version,
            "submission_method": hr_or_website,
            "date_sent": date_sent,
            "answer": "wait to answer",
            "answer_date": None
        }
        try:
            self.collection.insert_one(job_entry)
            print("New job entry added successfully.")
            return True
        except Exception as e:
            print(f"Failed to insert job entry: {e}")
            return False

    def get_avg_of_cv_versions(self):
        cv_a_count = self.collection.count_documents({"cv_version": "A"})
        cv_b_count = self.collection.count_documents({"cv_version": "B"})
        cv_a_positive = self.collection.count_documents({"cv_version": "A", "answer": "yes"})
        cv_b_positive = self.collection.count_documents({"cv_version": "B", "answer": "yes"})

        avg_cv_a = cv_a_positive / cv_a_count if cv_a_count > 0 else 0
        avg_cv_b = cv_b_positive / cv_b_count if cv_b_count > 0 else 0

        return avg_cv_a, avg_cv_b

    def plot_success_rates(self):
        avg_cv_a, avg_cv_b = self.get_avg_of_cv_versions()

        versions = ['CV A', 'CV B']
        success_rates = [avg_cv_a, avg_cv_b]

        plt.bar(versions, success_rates, color=['blue', 'green'])
        plt.xlabel('CV Versions')
        plt.ylabel('Success Rate')
        plt.title('Success Rates of CV Versions')
        plt.show()

    def compare_submission_methods(self):
        hr_count = self.collection.count_documents({"submission_method": "hr"})
        hr_positive = self.collection.count_documents({"submission_method": "hr", "answer": "yes"})
        website_count = self.collection.count_documents({"submission_method": "website"})
        website_positive = self.collection.count_documents({"submission_method": "website", "answer": "yes"})

        hr_success_rate = hr_positive / hr_count if hr_count > 0 else 0
        website_success_rate = website_positive / website_count if website_count > 0 else 0

        return hr_success_rate, website_success_rate

    def analyze_preference_by_submission_method(self):
        hr_cv_a_positive = self.collection.count_documents(
            {"submission_method": "hr", "cv_version": "A", "answer": "yes"})
        hr_cv_b_positive = self.collection.count_documents(
            {"submission_method": "hr", "cv_version": "B", "answer": "yes"})
        website_cv_a_positive = self.collection.count_documents(
            {"submission_method": "website", "cv_version": "A", "answer": "yes"})
        website_cv_b_positive = self.collection.count_documents(
            {"submission_method": "website", "cv_version": "B", "answer": "yes"})

        return {
            'HR': {'CV A': hr_cv_a_positive, 'CV B': hr_cv_b_positive},
            'Website': {'CV A': website_cv_a_positive, 'CV B': website_cv_b_positive}
        }

    def get_company_names_by_answer(self):
        positive_companies = self.collection.distinct("company_name", {"answer": "yes"})
        negative_companies = self.collection.distinct("company_name", {"answer": {"$ne": "yes"}})

        return positive_companies, negative_companies

    def delete_jobs_before_2024(self):
        delete_result = self.collection.delete_many({"date_sent": {"$lt": "2024-01-01"}})
        print(f"Deleted {delete_result.deleted_count} jobs with dates before 2024.")

    def edit_line_by_company_name(self, company_name, link=None, title=None, cv_version=None, hr_or_website=None):
        update_fields = {}
        if link is not None:
            update_fields["job_link"] = link
        if title is not None:
            update_fields["job_title"] = title
        if cv_version is not None:
            update_fields["cv_version"] = cv_version
        if hr_or_website is not None:
            update_fields["submission_method"] = hr_or_website
        if not update_fields:
            print("No updates provided.")
            return

        update_result = self.collection.update_one(
            {"company_name": company_name},
            {"$set": update_fields}
        )
        if update_result.matched_count > 0:
            print(f"Line for company '{company_name}' updated successfully.")
        else:
            print(f"Company '{company_name}' not found.")

    def delete_lines_by_company_name(self, company_name):
        delete_result = self.collection.delete_many({"company_name": company_name})
        print(f"Deleted {delete_result.deleted_count} lines with company name '{company_name}'.")

    def does_hr_have_job_link(self, hr_name, job_link):
        # Get all job links associated with the given HR name using the existing method
        job_links = self.get_job_links_by_hr_name(hr_name)

        # Check if the specific job link exists in the list of job links
        return job_link in job_links

    def does_hr_have_job(self, hr_name, company_name, job_title):
        # Query the collection to find if an entry exists with the specified HR name, company name, and job title
        query = {
            "submission_method": {"$regex": f".*{hr_name}.*", "$options": "i"},
            "company_name": company_name,
            "job_title": job_title
        }

        # Check if a document matching the criteria exists
        return self.collection.count_documents(query) > 0

    def get_job_links_by_hr_name(self, hr_name):
        # Query the collection for job entries with the HR name in the submission method
        jobs = self.collection.find({"submission_method": {"$regex": f".*{hr_name}.*", "$options": "i"}})

        # Extract job links from the query results
        job_links = [job["job_link"] for job in jobs if "job_link" in job]

        return job_links

    def get_jobs_by_company_name(self, company_name):
        try:
            jobs = list(self.collection.find({"company_name": company_name}))
            if jobs:
                print(f"Found {len(jobs)} jobs for company '{company_name}'.")
                return jobs
            else:
                print(f"No jobs found for company '{company_name}'.")
                return []
        except Exception as e:
            print(f"An error occurred while retrieving jobs for company '{company_name}': {e}")
            return []

    def export_website_submitted_jobs_to_excel(self, output_file_path="WebsiteSubmittedJobs.xlsx"):
        try:
            # If the Excel file exists, load it; otherwise, create a new workbook
            if os.path.exists(output_file_path):
                workbook = load_workbook(output_file_path)
                sheet = workbook.active
            else:
                workbook = Workbook()
                sheet = workbook.active

                # Add headers to the worksheet (only when creating a new workbook)
                headers = ["Company Name", "Job Link", "Job Title", "CV Version", "Submission Method", "Date Sent",
                           "Answer", "Answer Date"]
                sheet.append(headers)

                # Apply yellow fill to the headers
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                for cell in sheet[1]:  # Apply the fill to the first row (headers)
                    cell.fill = yellow_fill

            # Retrieve jobs that were submitted via the website
            website_jobs = list(self.collection.find({"submission_method": {"$regex": "^web", "$options": "i"}}))

            if not website_jobs:
                print("No jobs found that were submitted via the website.")
                return

            # Add job data to the worksheet
            for job in website_jobs:
                next_row = sheet.max_row + 1
                job_data = (
                    job.get("company_name", ""),
                    job.get("job_link", ""),
                    job.get("job_title", ""),
                    job.get("cv_version", ""),
                    job.get("submission_method", ""),
                    job.get("date_sent", datetime.datetime.today().strftime('%Y-%m-%d')),
                    job.get("answer", "wait to answer"),
                    job.get("answer_date", "wait to answer")
                )
                for col_num, value in enumerate(job_data, start=1):
                    sheet.cell(row=next_row, column=col_num, value=value)

            # Save the workbook to the specified file path
            workbook.save(output_file_path)
            print(f"Website-submitted jobs exported successfully to {output_file_path}.")

        except Exception as e:
            print(f"An error occurred while exporting website-submitted jobs: {e}")

    def export_all_jobs_to_excel(self, output_file_path="JobsExport.xlsx"):
        try:
            # If the Excel file exists, load it; otherwise, create a new workbook
            if os.path.exists(output_file_path):
                workbook = load_workbook(output_file_path)
                sheet = workbook.active
            else:
                workbook = Workbook()
                sheet = workbook.active

                # Add headers to the worksheet (only when creating a new workbook)
                headers = ["Company Name", "Job Link", "Job Title", "CV Version", "Submission Method", "Date Sent",
                           "Answer", "Answer Date"]
                sheet.append(headers)

                # Apply yellow fill to the headers
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                for cell in sheet[1]:  # Apply the fill to the first row (headers)
                    cell.fill = yellow_fill

            # Retrieve jobs from MongoDB and add them to the worksheet
            jobs = list(self.collection.find())

            for job in jobs:
                next_row = sheet.max_row + 1
                job_data = (
                    job.get("company_name", ""),
                    job.get("job_link", ""),
                    job.get("job_title", ""),
                    job.get("cv_version", ""),
                    job.get("submission_method", ""),
                    job.get("date_sent", datetime.datetime.today().strftime('%Y-%m-%d')),
                    job.get("answer", "wait to answer"),
                    job.get("answer_date", "wait to answer")
                )
                for col_num, value in enumerate(job_data, start=1):
                    sheet.cell(row=next_row, column=col_num, value=value)

            # Save the workbook to the specified file path
            workbook.save(output_file_path)
            print(f"All jobs exported successfully to {output_file_path}.")

        except Exception as e:
            print(f"An error occurred while exporting jobs: {e}")

    def export_hr_submitted_jobs_to_excel(self, output_file_path="HRSubmittedJobs.xlsx"):
        try:
            # If the Excel file exists, load it; otherwise, create a new workbook
            if os.path.exists(output_file_path):
                workbook = load_workbook(output_file_path)
                sheet = workbook.active
            else:
                workbook = Workbook()
                sheet = workbook.active

                # Add headers to the worksheet (only when creating a new workbook)
                headers = ["Company Name", "Job Link", "Job Title", "CV Version", "Submission Method", "Date Sent",
                           "Answer", "Answer Date"]
                sheet.append(headers)

                # Apply yellow fill to the headers
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                for cell in sheet[1]:  # Apply the fill to the first row (headers)
                    cell.fill = yellow_fill

            # Retrieve jobs that were submitted via HR
            hr_jobs = list(self.collection.find({"submission_method": {"$regex": "^hr", "$options": "i"}}))

            if not hr_jobs:
                print("No jobs found that were submitted via HR.")
                return

            # Add job data to the worksheet
            for job in hr_jobs:
                next_row = sheet.max_row + 1
                job_data = (
                    job.get("company_name", ""),
                    job.get("job_link", ""),
                    job.get("job_title", ""),
                    job.get("cv_version", ""),
                    job.get("submission_method", ""),
                    job.get("date_sent", datetime.datetime.today().strftime('%Y-%m-%d')),
                    job.get("answer", "wait to answer"),
                    job.get("answer_date", "wait to answer")
                )
                for col_num, value in enumerate(job_data, start=1):
                    sheet.cell(row=next_row, column=col_num, value=value)

            # Save the workbook to the specified file path
            workbook.save(output_file_path)
            print(f"HR-submitted jobs exported successfully to {output_file_path}.")

        except Exception as e:
            print(f"An error occurred while exporting HR-submitted jobs: {e}")

    def set_job_answer(self, answer, company_name, title):
        try:
            # Get the current date for the answer date
            answer_date = datetime.datetime.today().strftime('%Y-%m-%d')

            # Update the job document with the provided answer and answer date
            update_result = self.collection.update_one(
                {"company_name": company_name, "job_title": title},
                {"$set": {"answer": answer, "answer_date": answer_date}}
            )

            if update_result.matched_count > 0:
                print(f"Job at '{company_name}' for title '{title}' updated with answer '{answer}'.")
            else:
                print(f"No job found for company '{company_name}' with title '{title}'.")

        except Exception as e:
            print(f"An error occurred while updating the job answer: {e}")
