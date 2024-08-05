# /Users/yrdnqldrwn/Desktop/SOFTWARE/PayChatm/Info_aboutCVsubmitted/CreateExcelFile.py
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


class CreateExcelFile:
    def __init__(self, filename=""):
        self.filename = filename
        self.data = {}

    def set_filename(self, filename):
        self.filename = filename if filename.endswith('.xlsx') else filename + '.xlsx'

    def get_filename(self):
        return self.filename

    def get_user_data(self):
        columns = input("Enter the column names separated by commas: ").split(',')
        num_rows = int(input("Enter the number of rows: "))

        for column in columns:
            self.data[column.strip()] = []

        for i in range(num_rows):
            print(f"\nEntering data for row {i + 1}:")
            for column in columns:
                value = input(f"Enter value for {column.strip()}: ")
                self.data[column.strip()].append(value)

    def save_to_excel(self):
        df = pd.DataFrame(self.data)

        # Create a new workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active

        # Write the column headers
        for col_num, column_title in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num, value=column_title)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write the data rows
        for row_num, row_data in enumerate(df.values, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=cell_value)

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # Save the workbook
        wb.save(self.filename)
        return True

    def create_excel(self, name=""):
        if len(name) < 1:
            return False

        self.filename = name + '.xlsx'

        try:
            self.get_user_data()
            return self.save_to_excel()
        except Exception as e:
            print(f"An error occurred while creating the Excel file: {e}")
            return False
