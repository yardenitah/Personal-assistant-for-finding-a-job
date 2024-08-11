# /Users/yrdnqldrwn/Desktop/SOFTWARE/PayChatm/Info_aboutCVsubmitted/ExcelManager.py
import os
from openpyxl import load_workbook, Workbook
import matplotlib.pyplot as plt
import datetime


class ExcelManager:
    def __init__(self, file_path=""):
        self.file_path = file_path

    def add_line(self, company_name, link, title, cv_version, hr_or_website):
        workbook = load_workbook(self.file_path)
        sheet = workbook.active

        next_row = sheet.max_row + 1
        date_sent = datetime.datetime.today().strftime('%Y-%m-%d')
        answer = "wait to answer"
        answer_date = "wait to answer"

        new_line = (company_name, link, title, cv_version, hr_or_website, date_sent, answer, answer_date)
        for col_num, value in enumerate(new_line, start=1):
            sheet.cell(row=next_row, column=col_num, value=value)

        workbook.save(self.file_path)
        print(f"New line added to {self.file_path}.")

    def get_avg_of_cv_versions(self):
        workbook = load_workbook(self.file_path)
        sheet = workbook.active

        cv_a_count = 0
        cv_b_count = 0
        cv_a_positive = 0
        cv_b_positive = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            cv_version = row[3]
            answer = row[6]

            if answer.lower() == "yes":
                if cv_version.lower() == "a":
                    cv_a_positive += 1
                    cv_a_count += 1
                elif cv_version.lower() == "b":
                    cv_b_positive += 1
                    cv_b_count += 1
            else:
                if cv_version.lower() == "a":
                    cv_a_count += 1
                elif cv_version.lower() == "b":
                    cv_b_count += 1

        avg_cv_a = cv_a_positive / cv_a_count if cv_a_count > 0 else 0
        avg_cv_b = cv_b_positive / cv_b_count if cv_b_count > 0 else 0

        return avg_cv_a, avg_cv_b

    def plot_success_rates(self):
        avg_cv_a, avg_cv_b = self.get_avg_of_cv_versions()

        versions = ['CV A', 'CV B']
        success_rates = [avg_cv_a, avg_cv_b]

        plt.bar(versions, success_rates, color=['blue', 'green'])
        plt.xlabel('CV Versions')
        plt.ylabel('Success Rate')
        plt.title('Success Rates of CV Versions')
        plt.show()

    def compare_submission_methods(self):
        workbook = load_workbook(self.file_path)
        sheet = workbook.active

        hr_count = 0
        hr_positive = 0
        website_count = 0
        website_positive = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            submission_method = row[4]
            answer = row[6]

            if answer.lower() == "yes":
                if submission_method.lower() == "hr":
                    hr_positive += 1
                    hr_count += 1
                elif submission_method.lower() == "website":
                    website_positive += 1
                    website_count += 1
            else:
                if submission_method.lower() == "hr":
                    hr_count += 1
                elif submission_method.lower() == "website":
                    website_count += 1

        hr_success_rate = hr_positive / hr_count if hr_count > 0 else 0
        website_success_rate = website_positive / website_count if website_count > 0 else 0

        return hr_success_rate, website_success_rate

    def analyze_preference_by_submission_method(self):
        workbook = load_workbook(self.file_path)
        sheet = workbook.active

        hr_cv_a_positive = 0
        hr_cv_b_positive = 0
        website_cv_a_positive = 0
        website_cv_b_positive = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            submission_method = row[4]
            cv_version = row[3]
            answer = row[6]

            if answer.lower() == "yes":
                if submission_method.lower() == "hr":
                    if cv_version.lower() == "a":
                        hr_cv_a_positive += 1
                    elif cv_version.lower() == "b":
                        hr_cv_b_positive += 1
                elif submission_method.lower() == "website":
                    if cv_version.lower() == "a":
                        website_cv_a_positive += 1
                    elif cv_version.lower() == "b":
                        website_cv_b_positive += 1

        return {
            'HR': {'CV A': hr_cv_a_positive, 'CV B': hr_cv_b_positive},
            'Website': {'CV A': website_cv_a_positive, 'CV B': website_cv_b_positive}
        }

    def get_company_names_by_answer(self):
        workbook = load_workbook(self.file_path)
        sheet = workbook.active

        positive_companies = []
        negative_companies = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            company_name = row[0]
            answer = row[6]

            if answer.lower() == "yes":
                positive_companies.append(company_name)
            else:
                negative_companies.append(company_name)

        return positive_companies, negative_companies

    def delete_jobs_before_2024(self):
        workbook = load_workbook(self.file_path)
        sheet = workbook.active

        rows_to_delete = []
        for row in sheet.iter_rows(min_row=2):
            date_sent = row[5].value
            if date_sent:
                try:
                    year = int(date_sent.split('-')[0])
                    if year < 2024:
                        rows_to_delete.append(row[0].row)
                except ValueError:
                    continue

        for row_num in reversed(rows_to_delete):
            sheet.delete_rows(row_num, 1)

        workbook.save(self.file_path)
        print(f"Deleted jobs with dates before 2024 from {self.file_path}.")

    def edit_line_by_company_name(self, company_name, link=None, title=None, cv_version=None, hr_or_website=None):
        workbook = load_workbook(self.file_path)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2):
            if row[0].value.lower() == company_name.lower():
                if link is not None:
                    row[1].value = link
                if title is not None:
                    row[2].value = title
                if cv_version is not None:
                    row[3].value = cv_version
                if hr_or_website is not None:
                    row[4].value = hr_or_website
                if row[5].value is None:
                    date_sent = datetime.datetime.today().strftime('%Y-%m-%d')
                    row[5].value = date_sent
                    # row[5].value = '2022-03-2'
                workbook.save(self.file_path)
                print(f"Line for company '{company_name}' updated successfully.")
                return

        print(f"Company '{company_name}' not found in the file.")

    def delete_lines_by_company_name(self, company_name):
        workbook = load_workbook(self.file_path)
        sheet = workbook.active

        rows_to_delete = []
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == None:
                break
            if row[0].value.lower() == company_name.lower():
                rows_to_delete.append(row[0].row)

        for row_num in reversed(rows_to_delete):
            sheet.delete_rows(row_num, 1)

        workbook.save(self.file_path)
        print(f"Deleted all lines with company name '{company_name}' from {self.file_path}.")
